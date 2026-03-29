"""
Generische Export-Mixins fuer Django REST Framework ViewSets.

Stellt wiederverwendbare XLSX-, PDF- und CSV-Streaming-Export-Aktionen
bereit, die in beliebige ViewSets integriert werden koennen.

Performance:
- CSV: Streaming-Response, kein Speicher-Overhead (empfohlen fuer grosse
  Datenmengen)
- XLSX: Openpyxl write-only mode fuer reduzierte Speichernutzung
- PDF: Chunked Queryset-Iteration (max 5000 Zeilen)
- Alle Exporte verwenden iterator() fuer speichereffiziente DB-Abfragen

Verwendung:
    class MyViewSet(ExportMixin, TenantViewSetMixin, ModelViewSet):
        export_fields = [
            {"key": "id", "label": "ID", "width": 10},
            {"key": "name", "label": "Name", "width": 30},
        ]
        export_filename = "mein_export"
        export_title = "Mein Export"
"""

import csv
import datetime
import io
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm, mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from rest_framework.decorators import action
from rest_framework.request import Request
from rest_framework.response import Response

from django.http import HttpResponse, StreamingHttpResponse

from drf_spectacular.utils import extend_schema

# Maximum rows for in-memory exports (XLSX, PDF)
MAX_EXPORT_ROWS_MEMORY = 5000
# Chunk size for iterator()
ITERATOR_CHUNK_SIZE = 500


class Echo:
    """Pseudo-buffer that returns the value written to it (for StreamingHttpResponse)."""

    def write(self, value):
        return value


class ExportMixin:
    """
    Mixin fuer ViewSets, das XLSX-, PDF- und CSV-Export-Aktionen bereitstellt.

    Konfiguration ueber Klassen-Attribute:
        export_fields: Liste von Dicts mit 'key', 'label', 'width' (optional)
        export_filename: Basis-Dateiname ohne Erweiterung
        export_title: Titel fuer den PDF-Export
    """

    export_fields: list[dict[str, Any]] = []
    export_filename: str = "export"
    export_title: str = "Export"

    def get_export_queryset(self, request: Request):
        """Liefert den gefilterten Queryset fuer den Export."""
        queryset = self.get_queryset()
        # Filter aus django-filters anwenden
        filterset_class = getattr(self, "filterset_class", None)
        if filterset_class:
            filterset = filterset_class(request.query_params, queryset=queryset)
            if filterset.is_valid():
                queryset = filterset.qs
        return queryset

    def get_export_fields(self) -> list[dict[str, Any]]:
        """Liefert die Export-Felder. Kann ueberschrieben werden."""
        return self.export_fields

    def get_export_filename(self) -> str:
        """Liefert den Dateinamen mit Datum."""
        today = datetime.date.today().strftime("%Y-%m-%d")
        return f"{self.export_filename}_{today}"

    def get_export_title(self) -> str:
        """Liefert den Titel fuer den Export."""
        return self.export_title

    def get_row_data(self, obj, fields: list[dict]) -> list[Any]:
        """
        Extrahiert die Daten einer Zeile aus einem Objekt.
        Unterstuetzt verschachtelte Attribute (z.B. 'student.first_name')
        und callable Attribute.
        """
        row = []
        for field in fields:
            key = field["key"]
            value = obj
            try:
                for part in key.split("."):
                    if value is None:
                        break
                    if isinstance(value, dict):
                        value = value.get(part, "")
                    else:
                        value = getattr(value, part, "")
                        if callable(value):
                            value = value()
            except (AttributeError, TypeError):
                value = ""

            # Formatierung
            if isinstance(value, datetime.datetime):
                value = value.strftime("%d.%m.%Y %H:%M")
            elif isinstance(value, datetime.date):
                value = value.strftime("%d.%m.%Y")
            elif isinstance(value, datetime.time):
                value = value.strftime("%H:%M")
            elif value is None:
                value = ""

            row.append(value)
        return row

    # ── CSV Streaming Export ────────────────────────────────────
    @extend_schema(
        summary="CSV-Export (Streaming)",
        description="Exportiert die gefilterten Daten als CSV-Datei mit Streaming fuer grosse Datenmengen.",
        responses={200: {"type": "string", "format": "binary"}},
    )
    @action(detail=False, methods=["get"], url_path="export-csv")
    def export_csv(self, request: Request) -> StreamingHttpResponse:
        """
        Streaming CSV export.

        Uses StreamingHttpResponse so that rows are written to the
        client as they are read from the database.  Memory usage is
        constant regardless of the number of rows.
        """
        queryset = self.get_export_queryset(request)
        fields = self.get_export_fields()
        filename = self.get_export_filename()

        pseudo_buffer = Echo()
        writer = csv.writer(pseudo_buffer, delimiter=";", quoting=csv.QUOTE_MINIMAL)

        def rows():
            # Header
            yield writer.writerow([f["label"] for f in fields])
            # Data rows – use iterator() for constant memory
            for obj in queryset.iterator(chunk_size=ITERATOR_CHUNK_SIZE):
                yield writer.writerow(
                    [str(v) for v in self.get_row_data(obj, fields)]
                )

        response = StreamingHttpResponse(rows(), content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
        return response

    # ── XLSX Export ──────────────────────────────────────────────

    @extend_schema(
        summary="XLSX-Export",
        description="Exportiert die gefilterten Daten als XLSX-Datei.",
        responses={200: {"type": "string", "format": "binary"}},
    )
    @action(detail=False, methods=["get"], url_path="export-xlsx")
    def export_xlsx(self, request: Request) -> HttpResponse:
        """Exportiert die Daten als XLSX-Datei."""
        queryset = self.get_export_queryset(request)
        fields = self.get_export_fields()
        filename = self.get_export_filename()
        title = self.get_export_title()

        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet(title=title[:31])

        # Styles
        header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(
            start_color="1F4E79", end_color="1F4E79", fill_type="solid"
        )
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell_font = Font(name="Calibri", size=10)
        cell_alignment = Alignment(vertical="top", wrap_text=True)

        # Header row
        header_cells = []
        for field in fields:
            cell = openpyxl.cell.WriteOnlyCell(ws, value=field["label"])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            header_cells.append(cell)
        ws.append(header_cells)

        # Data rows – use iterator() for constant memory
        row_count = 0
        for obj in queryset.iterator(chunk_size=ITERATOR_CHUNK_SIZE):
            row_data = self.get_row_data(obj, fields)
            data_cells = []
            for value in row_data:
                cell = openpyxl.cell.WriteOnlyCell(ws, value=value)
                cell.font = cell_font
                cell.alignment = cell_alignment
                data_cells.append(cell)
            ws.append(data_cells)
            row_count += 1
            if row_count >= MAX_EXPORT_ROWS_MEMORY:
                # Safety limit for in-memory exports
                break

        # Response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
        return response

    # ── PDF Export ───────────────────────────────────────────────

    @extend_schema(
        summary="PDF-Export",
        description="Exportiert die gefilterten Daten als PDF-Datei.",
        responses={200: {"type": "string", "format": "binary"}},
    )
    @action(detail=False, methods=["get"], url_path="export-pdf")
    def export_pdf(self, request: Request) -> HttpResponse:
        """Exportiert die Daten als PDF-Datei."""
        queryset = self.get_export_queryset(request)
        fields = self.get_export_fields()
        filename = self.get_export_filename()
        title = self.get_export_title()

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=1.5 * cm,
            rightMargin=1.5 * cm,
            topMargin=2 * cm,
            bottomMargin=1.5 * cm,
        )

        styles = getSampleStyleSheet()
        elements = []

        # Titel
        title_style = ParagraphStyle(
            "ExportTitle",
            parent=styles["Heading1"],
            fontSize=16,
            spaceAfter=6 * mm,
            alignment=1,  # Center
        )
        elements.append(Paragraph(title, title_style))

        # Datum
        date_style = ParagraphStyle(
            "ExportDate",
            parent=styles["Normal"],
            fontSize=9,
            textColor=colors.grey,
            alignment=1,
            spaceAfter=8 * mm,
        )
        elements.append(
            Paragraph(
                f"Erstellt am {datetime.date.today().strftime('%d.%m.%Y')}",
                date_style,
            )
        )

        # Tabellen-Daten
        cell_style = ParagraphStyle(
            "CellStyle",
            parent=styles["Normal"],
            fontSize=8,
            leading=10,
        )
        header_style = ParagraphStyle(
            "HeaderStyle",
            parent=styles["Normal"],
            fontSize=8,
            leading=10,
            textColor=colors.white,
            fontName="Helvetica-Bold",
        )

        # Header
        table_data = [
            [Paragraph(f["label"], header_style) for f in fields]
        ]

        # Daten-Zeilen – use iterator() with safety limit
        row_count = 0
        for obj in queryset.iterator(chunk_size=ITERATOR_CHUNK_SIZE):
            row_data = self.get_row_data(obj, fields)
            table_data.append(
                [Paragraph(str(v), cell_style) for v in row_data]
            )
            row_count += 1
            if row_count >= MAX_EXPORT_ROWS_MEMORY:
                break

        if len(table_data) <= 1:
            elements.append(
                Paragraph("Keine Daten vorhanden.", styles["Normal"])
            )
        else:
            # Spaltenbreiten berechnen
            available_width = landscape(A4)[0] - 3 * cm
            total_weight = sum(f.get("width", 15) for f in fields)
            col_widths = [
                (f.get("width", 15) / total_weight) * available_width
                for f in fields
            ]

            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            table.setStyle(
                TableStyle(
                    [
                        # Header
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 8),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                        ("TOPPADDING", (0, 0), (-1, 0), 6),
                        # Daten
                        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 1), (-1, -1), 8),
                        ("TOPPADDING", (0, 1), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
                        # Rahmen
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        # Zebra-Streifen
                        *[
                            (
                                "BACKGROUND",
                                (0, i),
                                (-1, i),
                                colors.HexColor("#F0F4F8"),
                            )
                            for i in range(2, len(table_data), 2)
                        ],
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]
                )
            )
            elements.append(table)

        # Footer mit Anzahl
        elements.append(Spacer(1, 8 * mm))
        footer_style = ParagraphStyle(
            "FooterStyle",
            parent=styles["Normal"],
            fontSize=8,
            textColor=colors.grey,
        )
        count = len(table_data) - 1
        truncated_note = (
            f" (limitiert auf {MAX_EXPORT_ROWS_MEMORY} Eintraege, verwenden Sie CSV fuer den vollstaendigen Export)"
            if row_count >= MAX_EXPORT_ROWS_MEMORY
            else ""
        )
        elements.append(
            Paragraph(
                f"Gesamt: {count} Eintr\u00e4ge{truncated_note}",
                footer_style,
            )
        )

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}.pdf"'
        return response
