"""
Export views for CSV and Excel downloads.

Provides export endpoints for:
- Transactions (CSV/Excel)
- Time Entries (CSV/Excel)
- Leave Requests (CSV/Excel)
"""

import csv
import io
import logging
from datetime import date

from django.http import HttpResponse
from drf_spectacular.utils import OpenApiParameter, extend_schema
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView

from core.permissions import IsLocationManagerOrAbove

logger = logging.getLogger("kassenbuch.export")


# ---------------------------------------------------------------------------
# Base Export Mixin
# ---------------------------------------------------------------------------


class ExportMixin:
    """Mixin providing CSV and Excel export utilities."""

    def generate_csv_response(self, filename, headers, rows):
        """Generate a CSV HTTP response."""
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response.write("\ufeff")  # BOM for Excel UTF-8 compatibility

        writer = csv.writer(response, delimiter=";")
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)

        return response

    def generate_excel_response(self, filename, sheet_name, headers, rows):
        """Generate an Excel HTTP response."""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
        header_alignment = Alignment(horizontal="center")

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-width columns
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row_idx in range(2, len(rows) + 2):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = min(max_length + 2, 50)

        # Write to response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())

        return response


# ---------------------------------------------------------------------------
# Transaction Export
# ---------------------------------------------------------------------------


class TransactionExportView(ExportMixin, APIView):
    """Export transactions as CSV or Excel."""

    permission_classes = [IsAuthenticated, IsLocationManagerOrAbove]

    @extend_schema(
        summary="Transaktionen exportieren",
        description="Exportiert Transaktionen als CSV oder Excel-Datei.",
        parameters=[
            OpenApiParameter(name="export_format", description="Export-Format", enum=["csv", "xlsx"], default="csv"),
            OpenApiParameter(name="start_date", description="Startdatum (YYYY-MM-DD)", type=str),
            OpenApiParameter(name="end_date", description="Enddatum (YYYY-MM-DD)", type=str),
            OpenApiParameter(name="status", description="Status-Filter", type=str),
            OpenApiParameter(name="group_id", description="Gruppen-Filter", type=int),
        ],
        tags=["Export"],
    )
    def get(self, request):
        """Export transactions."""
        from finance.models import Transaction

        export_format = request.query_params.get("export_format", "csv")
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        status_filter = request.query_params.get("status")
        group_id = request.query_params.get("group_id")

        # Build queryset
        queryset = Transaction.objects.select_related(
            "category", "group", "created_by", "approved_by"
        ).filter(is_deleted=False)

        # Apply user location filter (non-superadmin)
        if request.user.role != "super_admin" and request.user.location:
            queryset = queryset.filter(group__location=request.user.location)

        # Apply filters
        if start_date:
            queryset = queryset.filter(transaction_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(transaction_date__lte=end_date)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if group_id:
            queryset = queryset.filter(group_id=group_id)

        queryset = queryset.order_by("-transaction_date", "-created_at")

        # Build data
        headers = [
            "ID", "Datum", "Typ", "Kategorie", "Gruppe", "Betrag (\u20ac)",
            "Beschreibung", "Status", "Erstellt von", "Genehmigt von",
            "Erstellt am",
        ]

        rows = []
        for t in queryset:
            rows.append([
                t.id,
                t.transaction_date.strftime("%d.%m.%Y") if t.transaction_date else "",
                "Einnahme" if t.transaction_type == "income" else "Ausgabe",
                t.category.name if t.category else "",
                t.group.name if t.group else "",
                f"{t.amount:.2f}",
                t.description or "",
                t.get_status_display() if hasattr(t, "get_status_display") else t.status,
                t.created_by.get_full_name() if t.created_by else "",
                t.approved_by.get_full_name() if t.approved_by else "",
                t.created_at.strftime("%d.%m.%Y %H:%M") if t.created_at else "",
            ])

        today = date.today().strftime("%Y%m%d")

        if export_format == "xlsx":
            return self.generate_excel_response(
                f"transaktionen_{today}.xlsx", "Transaktionen", headers, rows
            )
        else:
            return self.generate_csv_response(
                f"transaktionen_{today}.csv", headers, rows
            )


# ---------------------------------------------------------------------------
# TimeEntry Export
# ---------------------------------------------------------------------------


class TimeEntryExportView(ExportMixin, APIView):
    """Export time entries as CSV or Excel."""

    permission_classes = [IsAuthenticated, IsLocationManagerOrAbove]

    @extend_schema(
        summary="Zeiteinträge exportieren",
        description="Exportiert Zeiteinträge als CSV oder Excel-Datei.",
        parameters=[
            OpenApiParameter(name="export_format", description="Export-Format", enum=["csv", "xlsx"], default="csv"),
            OpenApiParameter(name="start_date", description="Startdatum (YYYY-MM-DD)", type=str),
            OpenApiParameter(name="end_date", description="Enddatum (YYYY-MM-DD)", type=str),
            OpenApiParameter(name="user_id", description="Benutzer-Filter", type=int),
            OpenApiParameter(name="group_id", description="Gruppen-Filter", type=int),
        ],
        tags=["Export"],
    )
    def get(self, request):
        """Export time entries."""
        from timetracking.models import TimeEntry

        export_format = request.query_params.get("export_format", "csv")
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        user_id = request.query_params.get("user_id")
        group_id = request.query_params.get("group_id")

        # Build queryset
        queryset = TimeEntry.objects.select_related("user", "group").filter(is_deleted=False)

        # Apply user location filter
        if request.user.role != "super_admin" and request.user.location:
            queryset = queryset.filter(group__location=request.user.location)

        # Apply filters
        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__lte=end_date)
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        if group_id:
            queryset = queryset.filter(group_id=group_id)

        queryset = queryset.order_by("-date", "-start_time")

        # Build data
        headers = [
            "ID", "Datum", "Mitarbeiter:in", "Gruppe", "Beginn", "Ende",
            "Dauer (Min)", "Dauer (Std)", "Notizen",
        ]

        rows = []
        for entry in queryset:
            hours = entry.duration_minutes / 60 if entry.duration_minutes else 0
            rows.append([
                entry.id,
                entry.date.strftime("%d.%m.%Y") if entry.date else "",
                entry.user.get_full_name() if entry.user else "",
                entry.group.name if entry.group else "",
                entry.start_time.strftime("%H:%M") if entry.start_time else "",
                entry.end_time.strftime("%H:%M") if entry.end_time else "",
                entry.duration_minutes or 0,
                f"{hours:.1f}",
                entry.notes or "",
            ])

        today = date.today().strftime("%Y%m%d")

        if export_format == "xlsx":
            return self.generate_excel_response(
                f"zeiteintraege_{today}.xlsx", "Zeiteinträge", headers, rows
            )
        else:
            return self.generate_csv_response(
                f"zeiteintraege_{today}.csv", headers, rows
            )


# ---------------------------------------------------------------------------
# LeaveRequest Export
# ---------------------------------------------------------------------------


class LeaveRequestExportView(ExportMixin, APIView):
    """Export leave requests as CSV or Excel."""

    permission_classes = [IsAuthenticated, IsLocationManagerOrAbove]

    @extend_schema(
        summary="Urlaubsanträge exportieren",
        description="Exportiert Urlaubsanträge als CSV oder Excel-Datei.",
        parameters=[
            OpenApiParameter(name="export_format", description="Export-Format", enum=["csv", "xlsx"], default="csv"),
            OpenApiParameter(name="start_date", description="Startdatum (YYYY-MM-DD)", type=str),
            OpenApiParameter(name="end_date", description="Enddatum (YYYY-MM-DD)", type=str),
            OpenApiParameter(name="status", description="Status-Filter", type=str),
        ],
        tags=["Export"],
    )
    def get(self, request):
        """Export leave requests."""
        from timetracking.models import LeaveRequest

        export_format = request.query_params.get("export_format", "csv")
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        status_filter = request.query_params.get("status")

        # Build queryset
        queryset = LeaveRequest.objects.select_related(
            "user", "leave_type", "approved_by"
        ).filter(is_deleted=False)

        # Apply user location filter
        if request.user.role != "super_admin" and request.user.location:
            queryset = queryset.filter(user__location=request.user.location)

        # Apply filters
        if start_date:
            queryset = queryset.filter(start_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(end_date__lte=end_date)
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        queryset = queryset.order_by("-start_date")

        # Build data
        headers = [
            "ID", "Mitarbeiter:in", "Abwesenheitstyp", "Von", "Bis",
            "Tage", "Status", "Grund", "Genehmigt von", "Anmerkung",
        ]

        rows = []
        for leave in queryset:
            rows.append([
                leave.id,
                leave.user.get_full_name() if leave.user else "",
                leave.leave_type.name if leave.leave_type else "",
                leave.start_date.strftime("%d.%m.%Y") if leave.start_date else "",
                leave.end_date.strftime("%d.%m.%Y") if leave.end_date else "",
                leave.total_days or 0,
                leave.get_status_display() if hasattr(leave, "get_status_display") else leave.status,
                leave.reason or "",
                leave.approved_by.get_full_name() if leave.approved_by else "",
                leave.approval_notes or "",
            ])

        today = date.today().strftime("%Y%m%d")

        if export_format == "xlsx":
            return self.generate_excel_response(
                f"urlaubsantraege_{today}.xlsx", "Urlaubsanträge", headers, rows
            )
        else:
            return self.generate_csv_response(
                f"urlaubsantraege_{today}.csv", headers, rows
            )
